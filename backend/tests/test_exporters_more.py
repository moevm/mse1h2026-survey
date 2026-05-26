import io

from openpyxl import load_workbook

from utils.export_excel import ExcelExporter, MultiSurveyExporter, _excel_value, _question_label
from utils.export_pdf import PDFExporter
from utils.visualization_functions import Answer, Survey, SurveyVisualizer


def _visualizer_with_all_question_types():
    survey = Survey(
        id="survey",
        title="Export survey",
        description="Description",
        questions=[
            {"id": "choice", "report_label": "Choice label", "title": "Choice", "type": "radio", "answers": ["A", "B", "C"]},
            {"id": "number", "title": "Number", "type": "scale"},
            {"id": "boolean", "title": "Boolean", "type": "boolean"},
            {"id": "text", "title": "Text", "type": "text"},
        ],
    )
    answers = [
        Answer(
            id="a1",
            survey_id="survey",
            group="3341",
            answers=[
                {"id_question": "choice", "value": "A"},
                {"id_question": "number", "value": 5},
                {"id_question": "boolean", "value": "true"},
                {"id_question": "text", "value": "alpha beta"},
            ],
        ),
        Answer(
            id="a2",
            survey_id="survey",
            group="3342",
            answers=[
                {"id_question": "choice", "value": "B"},
                {"id_question": "number", "value": 3},
                {"id_question": "boolean", "value": "false"},
                {"id_question": "text", "value": ""},
            ],
        ),
    ]
    return SurveyVisualizer([survey], answers)


def test_excel_helpers_cover_labels_and_complex_values():
    assert _question_label({"report_label": "Custom"}, 2) == "Custom"
    assert _question_label({}, 2) == "Question 2"
    assert _excel_value(None) == ""
    assert _excel_value(["A", "B"]) == "['A', 'B']"


def test_excel_exporter_writes_all_question_type_sheets():
    visualizer = _visualizer_with_all_question_types()

    data = ExcelExporter(visualizer).export_survey_to_excel("survey")
    workbook = load_workbook(io.BytesIO(data), read_only=True)

    assert workbook.sheetnames == ["General Information", "Raw Data", "Q1", "Q2", "Q3", "Q4"]
    assert workbook["Raw Data"]["A1"].value == "Answer ID"
    assert workbook["Q1"]["A1"].value == "Choice label: Choice"
    assert workbook["Q1"]["A4"].value == "Response Distribution"
    assert workbook["Q2"]["A4"].value == "Numeric Statistics"
    assert workbook["Q3"]["A4"].value == "Boolean Statistics"
    assert workbook["Q4"]["A4"].value == "Text Statistics"


def test_excel_exporter_handles_missing_survey_and_empty_data():
    visualizer = SurveyVisualizer([
        Survey(id="empty", title="Empty", description="", questions=[])
    ], [])

    data = ExcelExporter(visualizer).export_survey_to_excel("empty")
    workbook = load_workbook(io.BytesIO(data), read_only=True)

    assert workbook["Raw Data"]["A1"].value == "No data available"

    try:
        ExcelExporter(visualizer).export_survey_to_excel("missing")
    except ValueError as exc:
        assert "missing" in str(exc)
    else:
        raise AssertionError("Expected ValueError for missing survey")


def test_multi_survey_exporter_builds_table_of_contents():
    visualizer = _visualizer_with_all_question_types()

    data = MultiSurveyExporter(visualizer).export_all_surveys(filename=None)
    workbook = load_workbook(io.BytesIO(data), read_only=True)

    assert workbook.sheetnames[0] == "Table of Contents"
    assert workbook["Table of Contents"]["A1"].value == "All Surveys Report"
    assert workbook["Table of Contents"]["B7"].value == "Export survey"


def test_pdf_exporter_covers_question_type_branches_without_charts():
    visualizer = _visualizer_with_all_question_types()

    data = PDFExporter(visualizer).export_survey_to_pdf("survey", include_charts=False)

    assert data.startswith(b"%PDF")


def test_pdf_exporter_covers_chart_branches():
    visualizer = _visualizer_with_all_question_types()

    data = PDFExporter(visualizer).export_survey_to_pdf("survey", include_charts=True)

    assert data.startswith(b"%PDF")


def test_pdf_exporter_handles_missing_survey_and_empty_question_values():
    visualizer = SurveyVisualizer([
        Survey(id="empty", title="Empty", description="", questions=[{"id": "q", "title": "Q", "type": "text"}])
    ], [])
    exporter = PDFExporter(visualizer)

    assert exporter._create_question_analysis("empty", {"id": "q", "title": "Q", "type": "text"}, False, 1) == []

    try:
        exporter.export_survey_to_pdf("missing")
    except ValueError as exc:
        assert "missing" in str(exc)
    else:
        raise AssertionError("Expected ValueError for missing survey")
