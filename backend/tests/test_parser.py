import pandas as pd
from openpyxl import Workbook

from utils.parser import _parse_sheet


def _worksheet_and_dataframe(rows, merged_ranges=None):
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)

    for merged_range in merged_ranges or []:
        ws.merge_cells(merged_range)

    values = [[cell.value for cell in row] for row in ws.iter_rows()]
    return ws, pd.DataFrame(values)


def test_parse_sheet_keeps_merged_teacher_and_lesson_types():
    worksheet, dataframe = _worksheet_and_dataframe(
        [
            ["Преподаватель", "Дисциплина", "Лекции/лабы, практики", "Группа Программная инженерия"],
            ["Преподаватель 1", "Анализ, моделирование и оптимизация систем", "лекции", "1303, 1306"],
            [None, "Анализ, моделирование и оптимизация систем", "лабы", "1303, 1306"],
        ],
        merged_ranges=["A2:A3"],
    )

    records = _parse_sheet(dataframe, worksheet)

    assert records == [
        {
            "teacher": "Преподаватель 1",
            "discipline": "Анализ, моделирование и оптимизация систем (лекции)",
            "groups": ["1303", "1306"],
            "source_row": 2,
        },
        {
            "teacher": "Преподаватель 1",
            "discipline": "Анализ, моделирование и оптимизация систем (лабы)",
            "groups": ["1303", "1306"],
            "source_row": 3,
        },
    ]


def test_parse_sheet_skips_empty_rows_and_reads_all_group_columns():
    worksheet, dataframe = _worksheet_and_dataframe([
        ["noise", None, None, None, None],
        ["Преподаватель", "Дисциплина", "Лекции/лабы, практики", "Группа ПИ", "Группа ПМ"],
        [None, None, None, None, None],
        ["Преподаватель 2", "Базы данных", "практики", "3341, 3342", "3381"],
    ])

    records = _parse_sheet(dataframe, worksheet)

    assert records == [
        {
            "teacher": "Преподаватель 2",
            "discipline": "Базы данных (практики)",
            "groups": ["3341", "3342", "3381"],
            "source_row": 4,
        },
    ]
