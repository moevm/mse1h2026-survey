import io

from openpyxl import load_workbook

from models import Answer, Discipline, Group, GroupTeacherDiscipline, Survey, Teacher
from utils.export_excel import ExcelExporter
from utils.export_pdf import PDFExporter
from utils.report_factory import build_report_visualizer
from utils.visualization_functions import get_question_groups


def _create_blueprint_survey(db_session, group_name):
    survey = Survey(
        title="Report blueprint survey",
        description="Description",
        groups=[group_name],
        is_active=True,
        questions=[
            {"id": "plain", "title": "Plain question", "type": "text"},
            {
                "id": "bp",
                "title": "Blueprint",
                "type": "blueprint",
                "options": [
                    {"id": "subject", "title": "{{subject}} quality", "type": "text"},
                    {"id": "teacher", "title": "{{teacher}} teaches {{subject}}", "type": "text"},
                ],
            },
        ],
    )
    group = Group(name=group_name)
    teacher = Teacher(name="Teacher A")
    subject_lecture = Discipline(name="Databases (lectures)")
    subject_lab = Discipline(name="Databases (labs)")

    db_session.add_all([survey, group, teacher, subject_lecture, subject_lab])
    db_session.flush()
    db_session.add_all([
        GroupTeacherDiscipline(
            survey_id=survey.id,
            group_id=group.id,
            teacher_id=teacher.id,
            discipline_id=subject_lecture.id,
            source_order=1,
        ),
        GroupTeacherDiscipline(
            survey_id=survey.id,
            group_id=group.id,
            teacher_id=teacher.id,
            discipline_id=subject_lab.id,
            source_order=2,
        ),
    ])
    db_session.flush()
    return survey


def test_report_factory_expands_blueprints_for_answer_group(db_session):
    group_name = "r3341"
    survey = _create_blueprint_survey(db_session, group_name)
    answer = Answer(
        survey_id=survey.id,
        group=group_name,
        answers=[
            {"id_question": "plain", "value": "plain answer"},
            {"id_question": "bp-Databases (lectures)-subject", "value": "lecture subject answer"},
            {"id_question": "bp-Databases (lectures)-Teacher A-teacher", "value": "lecture teacher answer"},
            {"id_question": "bp-Databases (labs)-subject", "value": "lab subject answer"},
            {"id_question": "bp-Databases (labs)-Teacher A-teacher", "value": "lab teacher answer"},
        ],
    )
    db_session.add(answer)
    db_session.flush()

    visualizer = build_report_visualizer(survey, [answer], db_session)
    survey_id = str(survey.id)
    report_survey = visualizer.surveys[survey_id]
    titles = [question["title"] for question in report_survey.questions]

    assert titles == [
        "Plain question",
        "Databases (lectures) quality",
        "Teacher A teaches Databases (lectures)",
        "Databases (labs) quality",
        "Teacher A teaches Databases (labs)",
    ]
    groups = [
        (question.get("report_group_label"), question["title"])
        for question in report_survey.questions
        if question.get("report_group_label")
    ]
    assert groups == [
        ("Databases (lectures)", "Databases (lectures) quality"),
        ("Databases (lectures)", "Teacher A teaches Databases (lectures)"),
        ("Databases (labs)", "Databases (labs) quality"),
        ("Databases (labs)", "Teacher A teaches Databases (labs)"),
    ]
    data = visualizer.get_survey_data(survey_id).iloc[0]
    assert data["q_bp-Databases (lectures)-subject"] == "lecture subject answer"
    assert data["q_bp-Databases (labs)-Teacher A-teacher"] == "lab teacher answer"


def test_excel_and_pdf_exports_use_human_question_labels(db_session):
    group_name = "e3341"
    survey = _create_blueprint_survey(db_session, group_name)
    answer = Answer(
        survey_id=survey.id,
        group=group_name,
        answers=[{"id_question": "bp-Databases (lectures)-subject", "value": "ok"}],
    )
    db_session.add(answer)
    db_session.flush()
    visualizer = build_report_visualizer(survey, [answer], db_session)
    survey_id = str(survey.id)

    excel_bytes = ExcelExporter(visualizer).export_survey_to_excel(survey_id)
    workbook = load_workbook(io.BytesIO(excel_bytes), read_only=True)

    assert "Q2" in workbook.sheetnames
    assert workbook["Q2"]["A1"].value == "Question 2: Databases (lectures)"
    assert workbook["Q2"]["A2"].value == "Blueprint question group"
    assert workbook["Q2"]["A4"].value == "Subquestion 1: Databases (lectures) quality"
    assert str(survey.id) not in workbook["Q2"]["A1"].value

    pdf_bytes = PDFExporter(visualizer).export_survey_to_pdf(survey_id, include_charts=False)
    assert pdf_bytes.startswith(b"%PDF")


def test_pdf_blueprint_group_uses_subquestion_labels_and_charts(db_session):
    group_name = "p3341"
    survey = _create_blueprint_survey(db_session, group_name)
    answer = Answer(
        survey_id=survey.id,
        group=group_name,
        answers=[
            {"id_question": "bp-Databases (lectures)-subject", "value": "useful clear material"},
            {"id_question": "bp-Databases (lectures)-Teacher A-teacher", "value": "yes"},
        ],
    )
    db_session.add(answer)
    db_session.flush()

    visualizer = build_report_visualizer(survey, [answer], db_session)
    survey_id = str(survey.id)
    groups = get_question_groups(visualizer.surveys[survey_id].questions)
    blueprint_group = next(item for item in groups if item["type"] == "group")

    story = PDFExporter(visualizer)._create_question_group_analysis(
        survey_id,
        blueprint_group,
        include_charts=True,
        index=2,
    )
    paragraph_text = "\n".join(
        item.getPlainText()
        for item in story
        if hasattr(item, "getPlainText")
    )

    assert "Question 2: Databases (lectures)" in paragraph_text
    assert "Subquestion 1: Databases (lectures) quality" in paragraph_text
    assert "Subquestion 2: Teacher A teaches Databases (lectures)" in paragraph_text
    assert any(item.__class__.__name__ == "Image" for item in story)
