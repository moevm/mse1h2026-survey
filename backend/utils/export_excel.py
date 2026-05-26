import pandas as pd
from typing import List, Dict, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
import io

from utils.visualization_functions import (
    Survey,
    SurveyVisualizer,
    normalize_question_type,
    get_question_id,
    get_question_text,
    get_question_options,
    get_question_groups,
)


def _question_label(question: Dict, index: int) -> str:
    return question.get("report_label") or f"Question {index}"


def _excel_value(value):
    if value is None:
        return ""
    if isinstance(value, (list, dict, tuple, set)):
        return str(value)
    return str(value) if value.__class__.__name__ == "UUID" else value


class ExcelExporter:
    def __init__(self, visualizer: SurveyVisualizer):
        self.visualizer = visualizer

    def export_survey_to_excel(self, survey_id: int, filename: str = None) -> bytes:
        survey = self.visualizer.surveys.get(survey_id)
        if not survey:
            raise ValueError(f"Survey with ID {survey_id} not found")

        wb = Workbook()
        wb.remove(wb.active)

        self._create_info_sheet(wb, survey)
        self._create_raw_data_sheet(wb, survey_id, survey)

        for idx, item in enumerate(get_question_groups(survey.questions), start=1):
            if item["type"] == "group":
                self._create_question_group_stats_sheet(wb, survey_id, item, idx)
            else:
                self._create_question_stats_sheet(wb, survey_id, item["question"], idx)

        if filename:
            wb.save(filename)
            return None
        else:
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

    def _create_info_sheet(self, wb: Workbook, survey: Survey):
        ws = wb.create_sheet("General Information")

        title_font = Font(size=14, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws['A1'] = f"Survey Information: {survey.title}"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws.merge_cells('A1:D1')

        info_data = [
            ("Survey ID:", survey.id),
            ("Title:", survey.title),
            ("Description:", survey.description),
            ("Created:", survey.created_at.strftime("%d.%m.%Y %H:%M") if survey.created_at else "Not specified"),
            ("Questions:", len(survey.questions)),
            ("Responses:", len(self.visualizer.get_survey_data(survey.id))),
            ("Export date:", datetime.now().strftime("%d.%m.%Y %H:%M:%S"))
        ]

        for i, (label, value) in enumerate(info_data, start=3):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = header_font
            ws[f'B{i}'] = _excel_value(value)
            ws[f'A{i}'].border = border
            ws[f'B{i}'].border = border

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50

    def _create_raw_data_sheet(self, wb: Workbook, survey_id: int, survey: Survey):
        ws = wb.create_sheet("Raw Data")

        df = self.visualizer.get_survey_data(survey_id)

        if df.empty:
            ws['A1'] = "No data available"
            return

        headers = ['Answer ID', 'Group', 'Date']
        for q in survey.questions:
            headers.append(get_question_text(q))

        ws.append(headers)

        for idx, row in df.iterrows():
            row_data = [
                row.get('answer_id', ''),
                row.get('group', ''),
                row.get('created_at', '')
            ]
            for q in survey.questions:
                col_name = f'q_{get_question_id(q)}'
                row_data.append(row.get(col_name, ''))
            ws.append([_excel_value(value) for value in row_data])

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    def _create_question_stats_sheet(self, wb: Workbook, survey_id: int, question: Dict, index: int):
        sheet_name = f"Q{index}"
        ws = wb.create_sheet(sheet_name)
        self._write_question_stats(ws, survey_id, question, index, 1)

    def _create_question_group_stats_sheet(self, wb: Workbook, survey_id: int, group: Dict, index: int):
        sheet_name = f"Q{index}"
        ws = wb.create_sheet(sheet_name)

        ws['A1'] = f"Question {index}: {group['label']}"
        ws['A1'].font = Font(size=12, bold=True)
        ws.merge_cells('A1:F1')
        ws['A2'] = "Blueprint question group"
        ws.merge_cells('A2:F2')

        row = 4
        for sub_index, question in enumerate(group["questions"], start=1):
            row = self._write_question_stats(ws, survey_id, question, sub_index, row, prefix="Subquestion")
            row += 2

    def _write_question_stats(
        self,
        ws,
        survey_id: int,
        question: Dict,
        index: int,
        start_row: int,
        prefix: str = "Question",
    ) -> int:
        question_id = get_question_id(question)
        question_text = get_question_text(question)
        question_type = normalize_question_type(question.get('type', 'text'))
        question_label = question.get("report_label") or f"{prefix} {index}"

        ws[f'A{start_row}'] = f"{question_label}: {question_text}"
        ws[f'A{start_row}'].font = Font(size=12, bold=True)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)

        ws[f'A{start_row + 1}'] = f"Type: {question_type}"
        ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=6)

        values = self.visualizer.get_question_values(survey_id, question_id)

        row = start_row + 3

        if question_type == 'choice':
            stats = self.visualizer.stats.choice_stats(values, get_question_options(question))

            ws[f'A{row}'] = "Response Distribution"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = "Total responses"
            ws[f'B{row}'] = stats['total']
            row += 1
            ws[f'A{row}'] = "Unique options"
            ws[f'B{row}'] = stats['unique']
            row += 1
            ws[f'A{row}'] = "Most common"
            ws[f'B{row}'] = _excel_value(stats['most_common'])
            ws[f'C{row}'] = f"({stats['most_common_count']} times)"
            row += 2

            ws[f'A{row}'] = "Distribution:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            for option, count in stats['distribution'].items():
                percent = (count / stats['total'] * 100) if stats['total'] > 0 else 0
                ws[f'A{row}'] = _excel_value(option)
                ws[f'B{row}'] = count
                ws[f'C{row}'] = f"{percent:.1f}%"
                row += 1

        elif question_type in ['number', 'integer', 'float']:
            stats = self.visualizer.stats.numeric_stats(values)
            ws[f'A{row}'] = "Numeric Statistics"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            numeric_stats = [
                ("Total", stats.get('total', 0)),
                ("Mean", stats.get('mean', 0)),
                ("Median", stats.get('median', 0)),
                ("Std Dev", stats.get('std', 0)),
                ("Min", stats.get('min', 0)),
                ("Max", stats.get('max', 0)),
                ("Q1", stats.get('q1', 0)),
                ("Q3", stats.get('q3', 0))
            ]

            for label, value in numeric_stats:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1

        elif question_type == 'boolean':
            stats = self.visualizer.stats.boolean_stats(values)
            ws[f'A{row}'] = "Boolean Statistics"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = "Total"
            ws[f'B{row}'] = stats['total']
            row += 1
            ws[f'A{row}'] = "Yes"
            ws[f'B{row}'] = stats['true']
            if stats['total'] > 0:
                ws[f'C{row}'] = f"{stats['true_percent']:.1f}%"
            row += 1
            ws[f'A{row}'] = "No"
            ws[f'B{row}'] = stats['false']
            if stats['total'] > 0:
                false_percent = 100 - stats['true_percent'] - (stats['other']/stats['total']*100)
                ws[f'C{row}'] = f"{false_percent:.1f}%"
            row += 1
            ws[f'A{row}'] = "Other"
            ws[f'B{row}'] = stats['other']
            row += 1

        else:
            stats = self.visualizer.stats.text_stats(values)
            ws[f'A{row}'] = "Text Statistics"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = "Total"
            ws[f'B{row}'] = stats['total']
            row += 1
            ws[f'A{row}'] = "Filled"
            ws[f'B{row}'] = stats['filled']
            row += 1
            ws[f'A{row}'] = "Empty"
            ws[f'B{row}'] = stats['empty']
            row += 1
            ws[f'A{row}'] = "Avg length"
            ws[f'B{row}'] = stats['avg_length']
            row += 1
            ws[f'A{row}'] = "Unique"
            ws[f'B{row}'] = stats['unique']
            row += 2

            word_stats = self.visualizer.stats.word_frequency(values, top_n=20)
            if word_stats['top_words']:
                ws[f'A{row}'] = "Top 20 Words:"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

                ws[f'A{row}'] = "Word"
                ws[f'B{row}'] = "Frequency"
                row += 1

                for word, freq in word_stats['top_words'].items():
                    ws[f'A{row}'] = word
                    ws[f'B{row}'] = freq
                    row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        return row


class MultiSurveyExporter:
    def __init__(self, visualizer: SurveyVisualizer):
        self.visualizer = visualizer
        self.excel_exporter = ExcelExporter(visualizer)

    def export_all_surveys(self, filename: str = "all_surveys_report.xlsx") -> bytes:
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)

        toc_ws = wb.create_sheet("Table of Contents", 0)
        self._create_table_of_contents(toc_ws)

        for survey_id, survey in self.visualizer.surveys.items():
            temp_wb = Workbook()
            temp_exporter = ExcelExporter(self.visualizer)

            temp_exporter._create_info_sheet(temp_wb, survey)
            temp_exporter._create_raw_data_sheet(temp_wb, survey_id, survey)

            for idx, item in enumerate(get_question_groups(survey.questions), start=1):
                if item["type"] == "group":
                    temp_exporter._create_question_group_stats_sheet(temp_wb, survey_id, item, idx)
                else:
                    temp_exporter._create_question_stats_sheet(temp_wb, survey_id, item["question"], idx)

            for sheet_name in temp_wb.sheetnames:
                source_sheet = temp_wb[sheet_name]
                new_sheet_name = f"{survey.id}_{sheet_name}"[:31]
                target_sheet = wb.create_sheet(new_sheet_name)

                for row_idx, row in enumerate(source_sheet.iter_rows(), 1):
                    for col_idx, cell in enumerate(row, 1):
                        target_cell = target_sheet.cell(row=row_idx, column=col_idx)
                        target_cell.value = cell.value

                        if cell.has_style:
                            try:
                                if cell.font:
                                    target_cell.font = cell.font.copy()
                                if cell.fill:
                                    target_cell.fill = cell.fill.copy()
                                if cell.border:
                                    target_cell.border = cell.border.copy()
                                if cell.alignment:
                                    target_cell.alignment = cell.alignment.copy()
                            except:
                                pass

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        if filename:
            wb.save(filename)
            return None

        return output.getvalue()

    def _create_table_of_contents(self, ws):
        title_font = Font(size=14, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True)

        ws['A1'] = "All Surveys Report"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws.merge_cells('A1:D1')

        ws['A3'] = "Report date:"
        ws['B3'] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

        ws['A5'] = "Surveys:"
        ws['A5'].font = header_font

        ws['A6'] = "ID"
        ws['B6'] = "Title"
        ws['C6'] = "Questions"
        ws['D6'] = "Responses"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}6'].font = header_font
            ws[f'{col}6'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        row = 7
        for survey_id, survey in self.visualizer.surveys.items():
            answer_count = len(self.visualizer.get_survey_data(survey_id))
            ws[f'A{row}'] = _excel_value(survey_id)
            ws[f'B{row}'] = survey.title
            ws[f'C{row}'] = len(survey.questions)
            ws[f'D{row}'] = answer_count
            row += 1

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
