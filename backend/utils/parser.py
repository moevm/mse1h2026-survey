import io
import re
import time

import pandas as pd
import requests
from sqlalchemy.orm import Session
from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from models import Group, Teacher, Discipline, GroupTeacherDiscipline


def extract_sheet_id(url: str) -> str:
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not match:
        raise HTTPException(
            status_code=422,
            detail=f"Не удалось найти ID таблицы в ссылке: {url}",
        )
    return match.group(1)


def download_workbook(sheet_id: str) -> pd.ExcelFile:
    export_url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    )
    last_error = None
    for attempt in range(3):
        try:
            resp = requests.get(export_url, timeout=30)
            break
        except requests.RequestException as e:
            last_error = e
            if attempt < 2:
                time.sleep(1)
    else:
        raise HTTPException(status_code=503, detail=f"Ошибка сети при загрузке таблицы: {last_error}")

    if resp.status_code == 403:
        raise HTTPException(
            status_code=403,
            detail="Таблица закрыта. Откройте доступ 'Все, у кого есть ссылка' в Google Sheets.",
        )
    if resp.status_code != 200:
        raise HTTPException(
            status_code=502,
            detail=f"Google Sheets вернул HTTP {resp.status_code}",
        )

    workbook = pd.ExcelFile(io.BytesIO(resp.content), engine="openpyxl")
    workbook._merged_workbook = load_workbook(io.BytesIO(resp.content), read_only=False, data_only=True)
    return workbook


def _clean_name(value) -> str:
    if pd.isna(value):
        return ""
    s = str(value).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _parse_groups(cell_value) -> list[str]:
    if pd.isna(cell_value):
        return []
    groups = []
    for part in str(cell_value).split(","):
        g = re.sub(r"\(.*?\)", "", _clean_name(part)).strip()
        if g and g.replace(" ", "").isalnum():
            groups.append(g)
    return groups


def _get_merged_cell_value(worksheet, row_number: int, column_number: int):
    if worksheet is None:
        return ""

    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row_number <= max_row and min_col <= column_number <= max_col:
            return worksheet.cell(row=min_row, column=min_col).value

    return ""


def _find_header_indexes(headers: list[str], *needles: str) -> list[int]:
    return [
        idx
        for idx, header in enumerate(headers)
        if any(needle in header for needle in needles)
    ]


def _first_header_index(headers: list[str], fallback: int, *needles: str) -> int:
    indexes = _find_header_indexes(headers, *needles)
    return indexes[0] if indexes else fallback


def _detect_header_row(df: pd.DataFrame) -> int | None:
    for row_idx, row in df.head(10).iterrows():
        headers = [_clean_name(value).lower() for value in row.tolist()]
        has_teacher = bool(_find_header_indexes(headers, "преподав"))
        has_discipline = bool(_find_header_indexes(headers, "дисциплин", "предмет"))
        has_group = bool(_find_header_indexes(headers, "груп"))

        if has_teacher and has_discipline and has_group:
            return row_idx

    return None


def _parse_sheet(df: pd.DataFrame, worksheet=None) -> list[dict]:
    if df.empty:
        return []

    header_row_idx = _detect_header_row(df)
    if header_row_idx is None:
        return []

    headers = [_clean_name(value).lower() for value in df.iloc[header_row_idx].tolist()]
    if not any(headers):
        return []

    teacher_col = _first_header_index(headers, 0, "преподав")
    discipline_col = _first_header_index(headers, 1, "дисциплин", "предмет")
    lesson_type_col = _first_header_index(headers, 2, "лекц", "лаб", "практик")
    group_cols = _find_header_indexes(headers, "груп")

    if not group_cols:
        return []

    records = []
    last_discipline = ""

    for row_idx, row in df.iloc[header_row_idx + 1:].iterrows():
        teacher = _clean_name(row.iloc[teacher_col])
        discipline = _clean_name(row.iloc[discipline_col])

        if not teacher and not discipline:
            continue

        if not teacher:
            teacher = _clean_name(_get_merged_cell_value(worksheet, row_idx + 1, teacher_col + 1))

        if not discipline:
            discipline = _clean_name(_get_merged_cell_value(worksheet, row_idx + 1, discipline_col + 1))

        discipline = discipline or last_discipline

        if not teacher or not discipline:
            continue

        last_discipline = discipline

        lesson_type = _clean_name(row.iloc[lesson_type_col])
        discipline_full = f"{discipline} ({lesson_type})" if lesson_type else discipline

        groups = []
        for group_col in group_cols:
            groups += _parse_groups(row.iloc[group_col])

        if groups:
            records.append({
                "teacher": teacher,
                "discipline": discipline_full,
                "groups": groups,
                "source_row": row_idx + 1,
            })

    return records


def _fetch_records(url: str) -> list[dict]:
    sheet_id = extract_sheet_id(url)
    workbook = download_workbook(sheet_id)

    all_records: list[dict] = []
    merged_workbook = getattr(workbook, "_merged_workbook", None)
    for sheet_name in workbook.sheet_names:
        df = workbook.parse(sheet_name, header=None)
        worksheet = merged_workbook[sheet_name] if merged_workbook else None
        for record in _parse_sheet(df, worksheet):
            record["source_order"] = len(all_records)
            record["source_sheet"] = sheet_name
            all_records.append(record)

    if not all_records:
        raise HTTPException(
            status_code=422,
            detail=f"Не найдены листы в таблице или они пустые.",
        )

    return all_records


def _column_tag_name(name: str, fallback: str) -> str:
    lower_name = name.lower()
    if "преподав" in lower_name:
        return "teacher"
    if "груп" in lower_name:
        return "group"
    if "предмет" in lower_name or "дисциплин" in lower_name:
        return "subject"

    tag = re.sub(r"\s+", "_", name.strip())
    tag = re.sub(r"[{}]+", "", tag)
    return tag or fallback


def _get_sheet_columns(workbook: pd.ExcelFile) -> list[dict]:
    tag_counts: dict[str, int] = {}
    label_counts: dict[str, int] = {}
    columns = []

    for sheet_name in workbook.sheet_names:
        df = workbook.parse(sheet_name, header=None)
        if df.empty:
            continue

        header_row_idx = _detect_header_row(df)
        if header_row_idx is None:
            continue

        for idx, value in enumerate(df.iloc[header_row_idx].tolist(), start=1):
            label = _clean_name(value) or f"Колонка {idx}"
            base_tag = _column_tag_name(label, f"column_{idx}")
            tag_counts[base_tag] = tag_counts.get(base_tag, 0) + 1
            label_counts[label] = label_counts.get(label, 0) + 1
            suffix = f"_{tag_counts[base_tag]}" if tag_counts[base_tag] > 1 else ""
            visible_label = f"{label} ({label_counts[label]})" if label_counts[label] > 1 else label

            columns.append({
                "option": f"{sheet_name}: {visible_label}",
                "value": f"{{{{{base_tag}{suffix}}}}}",
                "sheet": sheet_name,
            })

    return columns


def get_sheet_columns(url: str) -> list[dict]:
    sheet_id = extract_sheet_id(url)
    workbook = download_workbook(sheet_id)

    if not workbook.sheet_names:
        raise HTTPException(status_code=422, detail="В таблице нет листов.")

    columns = _get_sheet_columns(workbook)
    if not columns:
        raise HTTPException(status_code=422, detail="В таблице не найдены столбцы.")

    return columns


def get_sheet_groups(url: str) -> list[str]:
    try:
        records = _fetch_records(url)
    except HTTPException:
        return []

    groups = {group for record in records for group in record.get("groups", [])}
    return sorted(groups)


def get_sheet_tags_for_group(url: str, group: str) -> list[str]:
    sheet_id = extract_sheet_id(url)
    workbook = download_workbook(sheet_id)

    if not workbook.sheet_names:
        return []

    columns = _get_sheet_columns(workbook)
    tags_by_sheet: dict[str, set[str]] = {}
    for column in columns:
        value = column["value"].strip("{}")
        tags_by_sheet.setdefault(column["sheet"], set()).add(value)

    allowed_tags: set[str] = set()
    merged_workbook = getattr(workbook, "_merged_workbook", None)
    for sheet_name in workbook.sheet_names:
        df = workbook.parse(sheet_name, header=None)
        worksheet = merged_workbook[sheet_name] if merged_workbook else None
        records = _parse_sheet(df, worksheet)
        if any(group in record.get("groups", []) for record in records):
            allowed_tags.update(tags_by_sheet.get(sheet_name, set()))

    return sorted(allowed_tags)


def _populate(records: list[dict], session: Session, survey_id: str = None) -> dict:
    teacher_cache: dict[str, Teacher] = {t.name: t for t in session.query(Teacher).all()}
    discipline_cache: dict[str, Discipline] = {d.name: d for d in session.query(Discipline).all()}
    group_cache: dict[str, Group] = {g.name: g for g in session.query(Group).all()}

    query = session.query(GroupTeacherDiscipline)
    if survey_id:
        query = query.filter(GroupTeacherDiscipline.survey_id == survey_id)
    else:
        query = query.filter(GroupTeacherDiscipline.survey_id.is_(None))

    existing_triples: set[tuple[int, int, int]] = {
        (a.group_id, a.teacher_id, a.discipline_id)
        for a in query.all()
    }

    def get_or_create_teacher(name: str) -> Teacher:
        if name not in teacher_cache:
            obj = Teacher(name=name)
            session.add(obj)
            session.flush()
            teacher_cache[name] = obj
        return teacher_cache[name]

    def get_or_create_discipline(name: str) -> Discipline:
        if name not in discipline_cache:
            obj = Discipline(name=name)
            session.add(obj)
            session.flush()
            discipline_cache[name] = obj
        return discipline_cache[name]

    def get_or_create_group(name: str) -> Group:
        if name not in group_cache:
            obj = Group(name=name)
            session.add(obj)
            session.flush()
            group_cache[name] = obj
        return group_cache[name]

    for rec in records:
        teacher = get_or_create_teacher(rec["teacher"])
        discipline = get_or_create_discipline(rec["discipline"])

        for group_name in rec["groups"]:
            group = get_or_create_group(group_name)

            triple = (group.id, teacher.id, discipline.id)
            if triple not in existing_triples:
                session.add(GroupTeacherDiscipline(
                    survey_id=survey_id,
                    group=group,
                    teacher=teacher,
                    discipline=discipline,
                    source_order=rec.get("source_order"),
                ))
                existing_triples.add(triple)
            else:
                session.query(GroupTeacherDiscipline).filter(
                    GroupTeacherDiscipline.survey_id == survey_id,
                    GroupTeacherDiscipline.group_id == group.id,
                    GroupTeacherDiscipline.teacher_id == teacher.id,
                    GroupTeacherDiscipline.discipline_id == discipline.id,
                ).update({"source_order": rec.get("source_order")})

    session.commit()

    return {
        "groups": len(group_cache),
        "teachers": len(teacher_cache),
        "disciplines": len(discipline_cache),
        "assignments": len(existing_triples),
    }


def parse_and_populate(url: str, session: Session, survey_id: str = None) -> dict:
    records = _fetch_records(url)
    stats = _populate(records, session, survey_id)
    return stats
